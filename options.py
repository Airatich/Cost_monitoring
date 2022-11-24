import pandas as pd
import openpyxl
# Вставь свой user agent внутри кавычек
# user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36'
user_agent="Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36"
path_to_project="/Users/airats/Documents/Projects/Cost_monitoring"

#Города, которые выгружаем
# cities_start = {"Moscow": "Москва", 'Saint Petersburg': 'Санкт-Петербург','Yekaterinburg': 'Екатеринбург',
#     'Nizhny Novgorod': 'Нижний Новгород', 'Novosibirsk': "Новосибирск", "Rostov-on-Don": "Ростов-на-Дону"}
# cities_start=['Москва',"Санкт-Петербург","Екатеринбург","Нижний Новгород","Новосибирск","Ростов-на-Дону"]
# print(cities_start)
# print(type(cities_start))
# # Введи толщины для Лист_рулон_гк
# # thickness_prod1=[4,6,8,10,12,16,18,20,28,32]
# # Введи толщины для Лист_рулон_хкßß
# thickness_prod2=[0.5,0.7,0.8,1,1.2,1.5,2]
# # Введи толщины для Лист_рулон_оцинк
# thickness_prod3=[0.4, 0.5, 0.55, 0.7, 1]
# # Введи толщины для Лист_рулон_окраш
# thickness_prod4=[0.4, 0.45, 0.5, 0.55, 0.7, 1]

### Заполняет толщины для Лист_рулон_гк
file=openpyxl.open("/Users/airats/Documents/Projects/Cost_monitoring/settings_file.xlsx",read_only=True)
sheet=file.active

thickness_prod1=[]
thickness_prod2=[]
thickness_prod3=[]
thickness_prod4=[]
cities_start=[]

# Данные, которые выгружаем

# Заполняет толщины для Лист_рулон_гк
for row in sheet.iter_rows(min_row=2,min_col=0,max_col=1):
    thickness_prod1.append(row[0].value)
thickness_prod1=[i for i in thickness_prod1 if i is not None]
print(thickness_prod1)
# Заполняет толщины для Лист_рулон_хк
for row in sheet.iter_rows(min_row=2,min_col=2,max_col=3):
    thickness_prod2.append(row[0].value)
thickness_prod2=[i for i in thickness_prod2 if i is not None]
print(thickness_prod2)
# Заполняет толщины для Лист_рулон_оцинк
for row in sheet.iter_rows(min_row=2,min_col=3,max_col=4):
    thickness_prod3.append(row[0].value)
thickness_prod3=[i for i in thickness_prod3 if i is not None]
print(thickness_prod3)
# Заполняет толщины для Лист_рулон_окраш
for row in sheet.iter_rows(min_row=2,min_col=4,max_col=5):
    thickness_prod4.append(row[0].value)
thickness_prod4=[i for i in thickness_prod4 if i is not None]
print(thickness_prod4)
for row in sheet.iter_rows(min_row=2,min_col=5,max_col=6):
    cities_start.append(row[0].value)
cities_start=[i for i in cities_start if i is not None]
print(cities_start)

