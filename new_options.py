import pandas as pd
import openpyxl
import all_prod
file = openpyxl.open(r"C:\Users\ADMIN\OneDrive\Документы\pyProjects\Cost_monitoring\тест.xlsx", read_only=True)
sheet = file.sheetnames

# print(len(sheet))
count=1
for i in sheet:
    sheet=file[i]
    # print("Продукт: ")
    name_prod=sheet[2][0].value


    # print("Толщины: ")
    thinkness_prod=[]
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=3):
        thinkness_prod.append(row[0].value)
    thinkness_prod = [i for i in thinkness_prod if i is not None]
    # print(thinkness_prod)


    # print("Города: ")
    cities=[]
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
        cities.append(row[0].value)
    cities = [i for i in cities if i is not None]

    print(f"Продукт: {name_prod}")
    print(f"Толщины: {thinkness_prod}")
    print(f'Города: {cities}')

    all_prod.main_prod(name_prod,thinkness_prod,cities,count)
    count+=1


